"""
FastAPI server for Monthly Reporter (App2.0) + run history for the dashboard.

Run from repo root:
  PYTHONPATH=. uvicorn api.main:app --reload --port 8000

Vite proxies /api → 8000 (see dashboard/vite.config.ts).
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional

import requests as _http_requests
from pydantic import BaseModel

import logging as _logging

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse

_api_log = _logging.getLogger("resgro.api")

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env")
except ImportError:
    pass

# Optional: set on Cloud Run so JSON includes absolute report_url (avoids clients fetching resgro.ai).
PUBLIC_API_BASE_URL = os.environ.get("PUBLIC_API_BASE_URL", "").rstrip("/")

from agents.monthly_reporter.cloud_app.marketing_upload_layout import (  # noqa: E402
    write_marketing_csvs_to_work_dir,
)
from agents.monthly_reporter.cloud_app.resgro_runner import (  # noqa: E402
    ReportInputs,
    generate_monthly_report_bundle,
)
from agents.deepdive.agent import run as run_deepdive  # noqa: E402
from agents.marketingreco.agent import run as run_marketingreco  # noqa: E402
from agents.campaign_review.agent import run as run_campaign_review  # noqa: E402
from agents.campaign_review.agent import to_json_safe as campaign_review_to_json_safe  # noqa: E402
from agents.marketingreco.resgro_ads_excel import resgro_ads_upload_rows  # noqa: E402
from shared.config.settings import account_information_csv_path  # noqa: E402
from shared.utils.account_directory import load_account_operators_csv  # noqa: E402
from shared.data_session import (  # noqa: E402
    create_session,
    get_session,
    get_session_data_dir,
    ingest_csv_files,
    ingest_zip_files,
    list_sessions,
    load_datasets,
    record_agent_run,
)
from agents.data_agent.agent import run_manual as data_agent_run_manual  # noqa: E402
from agents.data_agent.agent import run_autopilot as data_agent_run_autopilot  # noqa: E402
from agents.boss_agent.agent import run as boss_agent_run  # noqa: E402

RUNS_BASE = ROOT / "data" / "runs" / "monthly_reporter"
RUNS_BASE.mkdir(parents=True, exist_ok=True)
INDEX_PATH = RUNS_BASE / "index.jsonl"

DD_RUNS_BASE = ROOT / "data" / "runs" / "deepdive"
DD_RUNS_BASE.mkdir(parents=True, exist_ok=True)
MRK_RUNS_BASE = ROOT / "data" / "runs" / "marketingreco"
MRK_RUNS_BASE.mkdir(parents=True, exist_ok=True)
CR_RUNS_BASE = ROOT / "data" / "runs" / "campaign_review"
CR_RUNS_BASE.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="ResgroAI API", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def root():
    return {
        "ok": True,
        "service": "resgro-agents-api",
        "message": "Agents API. Use /api/health, /api/sessions, /api/chat, etc.",
        "health": "/api/health",
    }


@app.get("/api/account-directory")
def get_account_directory():
    """
    Unique operators from ``Business Name (original)`` with DoorDash login/password for dashboard autofill.
    Configure path via ``ACCOUNT_INFORMATION_CSV`` (defaults to repo-root ``Account Information-McDonalds.csv``).
    """
    path = account_information_csv_path()
    operators, warning = load_account_operators_csv(path)
    return {
        "path": str(path),
        "operators": operators,
        "warning": warning,
    }


def _append_index(rec: dict) -> None:
    with INDEX_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, default=str) + "\n")


def _prepare_ads_rows_file(input_path: Path, work_dir: Path) -> Path:
    """
    Normalize Ads manual upload into a CSV consumed by browser automation.

    - CSV: returned as-is.
    - Excel: reads sheet named "Ads" (case-insensitive), writes extracted CSV.
    """
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return input_path

    if suffix not in (".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"):
        raise HTTPException(400, "ads_sheet_file must be .csv or an Excel file.")

    try:
        import pandas as pd
    except ImportError as exc:
        raise HTTPException(500, "pandas is required to read Excel ads_sheet_file.") from exc

    try:
        xl = pd.ExcelFile(input_path)
    except Exception as exc:
        raise HTTPException(400, f"Failed to read Excel file: {exc}") from exc

    ads_sheet_name = next((s for s in xl.sheet_names if s.strip().lower() == "ads"), None)
    if not ads_sheet_name:
        raise HTTPException(400, 'Excel file must contain a sheet named "Ads".')

    try:
        ads_df = pd.read_excel(xl, sheet_name=ads_sheet_name)
    except Exception as exc:
        raise HTTPException(400, f'Failed to read "Ads" sheet: {exc}') from exc

    out_csv = work_dir / f"{input_path.stem}__ads_sheet.csv"
    ads_df.to_csv(out_csv, index=False)
    return out_csv


def _write_marketingreco_campaigns_excel(path: Path, result: dict) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offers"
    # store_id = National when FINANCIAL_DETAILED mapping exists; doordash_store_id = Reporting / Day-Slot key.
    headers = [
        "Store ID",
        "DoorDash Store ID",
        "Store Name",
        "Minimum Subtotal",
        "Slot Tags",
        "Campaign Name",
        "Status",
    ]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = Font(bold=True)
    mappings = result.get("campaign_mappings") or []
    for r, m in enumerate(mappings, start=2):
        tags = m.get("slot_tags", [])
        tags_str = ",".join(str(t) for t in tags) if isinstance(tags, list) else str(tags or "")
        ws.cell(row=r, column=1, value=m.get("store_id", ""))
        ws.cell(row=r, column=2, value=m.get("doordash_store_id", ""))
        ws.cell(row=r, column=3, value=m.get("store_name", ""))
        ws.cell(row=r, column=4, value=m.get("min_subtotal", 0))
        ws.cell(row=r, column=5, value=tags_str)
        ws.cell(row=r, column=6, value=m.get("campaign_name", ""))
        ws.cell(row=r, column=7, value=m.get("status", "Pending"))

    ads_plan = result.get("ads_plan") or {}
    slot_table = ads_plan.get("slot_table") or []
    if slot_table:
        wss = wb.create_sheet("Ads slots")
        sh = [
            "Merchant store ID",
            "Store name",
            "Slot",
            "Orders",
            "Sales",
            "Net total",
            "Profitability %",
            "Ad placement",
            "Budget estimate",
            "Weekly budget",
        ]
        for idx, h in enumerate(sh, start=1):
            cell = wss.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, row in enumerate(slot_table, start=2):
            wss.cell(row=r, column=1, value=row.get("store_id"))
            wss.cell(row=r, column=2, value=row.get("store_name"))
            wss.cell(row=r, column=3, value=row.get("slot"))
            wss.cell(row=r, column=4, value=row.get("orders"))
            wss.cell(row=r, column=5, value=row.get("sales"))
            wss.cell(row=r, column=6, value=row.get("net_total"))
            wss.cell(row=r, column=7, value=row.get("profitability_pct"))
            wss.cell(row=r, column=8, value=row.get("ad_placement"))
            wss.cell(row=r, column=9, value=row.get("budget_estimate"))
            wss.cell(row=r, column=10, value=row.get("weekly_budget"))

    resgro_ads = resgro_ads_upload_rows(ads_plan)
    if resgro_ads:
        wsr = wb.create_sheet("Ads")
        rh = ["Merchant store ID", "Slots", "Bid strategy", "Budget", "Campaign Name"]
        for idx, h in enumerate(rh, start=1):
            cell = wsr.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, row in enumerate(resgro_ads, start=2):
            wsr.cell(row=r, column=1, value=row["store_id"])
            wsr.cell(row=r, column=2, value=row["slots"])
            wsr.cell(row=r, column=3, value=row["bid_strategy"])
            wsr.cell(row=r, column=4, value=row["budget"])
            wsr.cell(row=r, column=5, value=row["campaign_name"])

    campaigns = ads_plan.get("campaigns") or []
    if campaigns:
        wsa = wb.create_sheet("Ads planner")
        ah = [
            "store_id",
            "store_name",
            "day_of_week",
            "daypart",
            "tier",
            "priority_rank",
            "target_audience",
            "start_date",
            "end_date",
            "bid_strategy",
            "bid_amount",
            "bid_display",
            "budget_weight",
            "allocation_pct",
            "campaign_name",
            "rationale",
            "order_count",
            "avg_aov",
            "median_aov",
            "mode_basket",
            "avg_profitability",
            "profitability_pct",
            "ad_penetration",
            "composite_score",
        ]
        for idx, h in enumerate(ah, start=1):
            cell = wsa.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, c in enumerate(campaigns, start=2):
            m = c.get("metrics") or {}
            wsa.cell(row=r, column=1, value=c.get("store_id"))
            wsa.cell(row=r, column=2, value=c.get("store_name"))
            wsa.cell(row=r, column=3, value=c.get("day_of_week"))
            wsa.cell(row=r, column=4, value=c.get("daypart"))
            wsa.cell(row=r, column=5, value=c.get("tier"))
            wsa.cell(row=r, column=6, value=c.get("priority_rank"))
            wsa.cell(row=r, column=7, value=c.get("target_audience"))
            wsa.cell(row=r, column=8, value=c.get("start_date"))
            wsa.cell(row=r, column=9, value=c.get("end_date"))
            wsa.cell(row=r, column=10, value=c.get("bid_strategy"))
            wsa.cell(row=r, column=11, value=c.get("bid_amount"))
            wsa.cell(row=r, column=12, value=c.get("bid_display"))
            wsa.cell(row=r, column=13, value=c.get("budget_weight"))
            wsa.cell(row=r, column=14, value=c.get("allocation_pct"))
            wsa.cell(row=r, column=15, value=c.get("campaign_name"))
            wsa.cell(row=r, column=16, value=c.get("rationale"))
            wsa.cell(row=r, column=17, value=m.get("order_count"))
            wsa.cell(row=r, column=18, value=m.get("avg_aov"))
            wsa.cell(row=r, column=19, value=m.get("median_aov"))
            wsa.cell(row=r, column=20, value=m.get("mode_basket"))
            wsa.cell(row=r, column=21, value=m.get("avg_profitability"))
            wsa.cell(row=r, column=22, value=m.get("profitability_pct"))
            wsa.cell(row=r, column=23, value=m.get("ad_penetration"))
            wsa.cell(row=r, column=24, value=m.get("composite_score"))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _read_all_runs(limit: int = 200) -> list[dict]:
    if not INDEX_PATH.is_file():
        return []
    rows: list[dict] = []
    with INDEX_PATH.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return list(reversed(rows[-limit:]))


@app.get("/api/health")
def health() -> dict:
    return {"ok": True, "service": "resgro-agents-api"}


# ── GCS large uploads (bypass Cloud Run 32 MiB body limit) ─────────────────

from api import gcs_uploads as _gcs  # noqa: E402


class _GcsFileSignSpec(BaseModel):
    filename: str
    content_type: str = "application/octet-stream"
    size_bytes: int


class _GcsSignRequest(BaseModel):
    files: list[_GcsFileSignSpec]


class _GcsObjectRef(BaseModel):
    object_path: str
    filename: str


class _SessionFromGcsRequest(BaseModel):
    operator_id: str
    operator_name: str = ""
    date_range: str = ""
    objects: list[_GcsObjectRef]


class _MonthlyReporterFromGcsRequest(BaseModel):
    operator_id: str = ""
    operator_name: str = ""
    pre_range: str
    post_range: str
    excluded_dates: str = ""
    dd_store_ids: str = ""
    ue_store_ids: str = ""
    dd_object: Optional[_GcsObjectRef] = None
    ue_object: Optional[_GcsObjectRef] = None
    marketing_objects: list[_GcsObjectRef] = []


@app.get("/api/uploads/status")
def uploads_status() -> dict:
    return {
        "enabled": _gcs.uploads_enabled(),
        "bucket": _gcs.GCS_UPLOAD_BUCKET or None,
        "direct_upload_max_bytes": 30 * 1024 * 1024,
        "gcs_max_bytes_per_file": _gcs.MAX_UPLOAD_BYTES,
    }


@app.post("/api/uploads/sign")
def uploads_sign(req: _GcsSignRequest):
    if not req.files:
        raise HTTPException(400, "Provide at least one file to sign.")
    try:
        return _gcs.sign_files([f.model_dump() for f in req.files])
    except RuntimeError as exc:
        raise HTTPException(503, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:
        _api_log.exception("uploads_sign failed")
        raise HTTPException(500, f"Failed to create signed URLs: {exc}") from exc


@app.post("/api/sessions/from-gcs")
def post_session_from_gcs(req: _SessionFromGcsRequest):
    if not req.objects:
        raise HTTPException(400, "Provide at least one uploaded object.")
    if not _gcs.uploads_enabled():
        raise HTTPException(503, "GCS uploads are not configured on this server.")

    work = Path(tempfile.mkdtemp(prefix="session_gcs_"))
    gcs_paths: list[str] = []
    try:
        uploaded_zips, csv_pairs, gcs_paths = _gcs.materialize_objects(
            [o.model_dump() for o in req.objects],
            work,
        )
        if not uploaded_zips and not csv_pairs:
            raise HTTPException(400, "No files found in GCS for this upload.")

        result = data_agent_run_manual(
            operator_id=req.operator_id.strip(),
            operator_name=req.operator_name.strip(),
            zip_paths=uploaded_zips if uploaded_zips else None,
            csv_pairs=csv_pairs if csv_pairs else None,
            date_range=req.date_range.strip(),
        )
        return JSONResponse(result)
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(500, str(exc)) from exc
    finally:
        shutil.rmtree(work, ignore_errors=True)
        _gcs.delete_objects(gcs_paths)


def _monthly_reporter_response(
    run_id: str,
    bundle: dict,
    t0: datetime,
    operator_id: str,
    operator_name: str,
) -> JSONResponse:
    out_dir = RUNS_BASE / run_id
    out_dir.mkdir(parents=True, exist_ok=True)

    full_name = bundle["filename"]
    (out_dir / full_name).write_bytes(bundle["excel_bytes"])

    date_name = bundle.get("date_export_filename")
    if bundle.get("date_export_bytes") and date_name:
        (out_dir / date_name).write_bytes(bundle["date_export_bytes"])  # type: ignore[index]

    preview = {"tables": bundle.get("tables") or {}, "summary_text": bundle.get("summary_text")}
    (out_dir / "preview.json").write_text(json.dumps(preview, default=str), encoding="utf-8")

    duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
    meta = {
        "run_id": run_id,
        "agent": "monthly_reporter",
        "operator_id": operator_id.strip() or "—",
        "operator_name": operator_name.strip(),
        "status": "success",
        "started": t0.isoformat(),
        "duration_s": round(duration_s, 2),
        "summary_text": bundle.get("summary_text"),
        "full_report_filename": full_name,
        "date_export_filename": date_name if bundle.get("date_export_bytes") else None,
    }
    (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

    _append_index(
        {
            "id": run_id,
            "agent": "monthly_reporter",
            "operator": operator_id.strip() or operator_name.strip() or "—",
            "status": "success",
            "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
            "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
        }
    )

    return JSONResponse(
        {
            "run_id": run_id,
            "summary_text": bundle.get("summary_text"),
            "preview": preview,
            "downloads": {
                "full": f"/api/runs/{run_id}/download/full",
                "date": f"/api/runs/{run_id}/download/date"
                if bundle.get("date_export_bytes")
                else None,
            },
        }
    )


@app.post("/api/runs/monthly-reporter/from-gcs")
def post_monthly_reporter_from_gcs(req: _MonthlyReporterFromGcsRequest):
    if not _gcs.uploads_enabled():
        raise HTTPException(503, "GCS uploads are not configured on this server.")

    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"mr_gcs_{run_id[:8]}_"))
    gcs_paths: list[str] = []

    try:
        if req.dd_object:
            p = _gcs.download_object_to_path(req.dd_object.object_path, work / "dd-data.csv")
            gcs_paths.append(req.dd_object.object_path)
            if not p.stat().st_size:
                raise HTTPException(400, "DoorDash file is empty.")

        if req.ue_object:
            p = _gcs.download_object_to_path(req.ue_object.object_path, work / "ue-data.csv")
            gcs_paths.append(req.ue_object.object_path)
            if not p.stat().st_size:
                raise HTTPException(400, "UberEats file is empty.")

        mkt_pairs: list[tuple[str, bytes]] = []
        for mobj in req.marketing_objects:
            local = work / Path(mobj.filename).name
            _gcs.download_object_to_path(mobj.object_path, local)
            gcs_paths.append(mobj.object_path)
            raw = local.read_bytes()
            if raw:
                mkt_pairs.append((mobj.filename, raw))
        if mkt_pairs:
            write_marketing_csvs_to_work_dir(work, mkt_pairs)

        inputs = ReportInputs(
            pre_range=req.pre_range.strip(),
            post_range=req.post_range.strip(),
            excluded_dates_text=req.excluded_dates.strip(),
            operator_name=req.operator_name.strip(),
            dd_store_ids_text=req.dd_store_ids.strip(),
            ue_store_ids_text=req.ue_store_ids.strip(),
        )
        bundle = generate_monthly_report_bundle(inputs, data_root=work)
        return _monthly_reporter_response(
            run_id,
            bundle,
            t0,
            req.operator_id,
            req.operator_name,
        )
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except Exception as exc:
        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "monthly_reporter",
                "operator": req.operator_id.strip() or "—",
                "status": "failed",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s)}s",
                "error": str(exc),
            }
        )
        raise HTTPException(500, str(exc)) from exc
    finally:
        shutil.rmtree(work, ignore_errors=True)
        _gcs.delete_objects(gcs_paths)


def _count_run_statuses(runs: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for run in runs:
        status = str(run.get("status") or "unknown")
        counts[status] = counts.get(status, 0) + 1
    return counts


def _build_overview() -> dict:
    sessions = list_sessions(limit=200)
    runs = _read_all_runs()[:200]
    manual_sessions = [s for s in sessions if (s.get("mode") or "manual") == "manual"]
    auto_sessions = [s for s in sessions if (s.get("mode") or "") in {"autopilot", "auto"}]
    ready_sessions = [s for s in sessions if s.get("status") == "ready"]

    active_agents: set[str] = set()
    for session in sessions:
        for agent_name, run_meta in (session.get("agent_runs") or {}).items():
            if run_meta.get("status") in {"success", "completed", "running"}:
                active_agents.add(agent_name)
    for run in runs:
        agent = run.get("agent")
        if agent:
            active_agents.add(str(agent))

    return {
        "stats": {
            "active_agents": len(active_agents),
            "total_sessions": len(sessions),
            "ready_sessions": len(ready_sessions),
            "manual_sessions": len(manual_sessions),
            "auto_sessions": len(auto_sessions),
            "runs_total": len(runs),
            "run_status_counts": _count_run_statuses(runs),
        },
        "recent_sessions": sessions[:8],
        "recent_runs": runs[:8],
    }


# ── Data Session endpoints ──────────────────────────────────────────────────

@app.post("/api/sessions")
async def post_session(
    operator_id: str = Form(...),
    operator_name: str = Form(""),
    date_range: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    mode: str = Form("manual"),
    zip_files: Optional[List[UploadFile]] = File(None),
    csv_files: Optional[List[UploadFile]] = File(None),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    """Create a data session. Upload data once, then run any agent with the session_id."""
    try:
        mode_norm = mode.strip().lower()

        if mode_norm == "autopilot":
            if not doordash_email.strip() or not doordash_password:
                raise HTTPException(400, "Autopilot mode requires doordash_email and doordash_password.")
            import asyncio
            result = await asyncio.to_thread(
                data_agent_run_autopilot,
                operator_id=operator_id.strip(),
                operator_name=operator_name.strip(),
                doordash_email=doordash_email.strip(),
                doordash_password=doordash_password,
                date_range=date_range.strip(),
                start_date=start_date.strip(),
                end_date=end_date.strip(),
            )
            return JSONResponse(result)

        work = Path(tempfile.mkdtemp(prefix="session_"))
        try:
            uploaded_zips: list[Path] = []
            csv_pairs: list[tuple[str, bytes]] = []

            if zip_files:
                for uf in zip_files:
                    if not uf.filename:
                        continue
                    raw = await uf.read()
                    if not raw:
                        continue
                    p = work / Path(uf.filename).name
                    p.write_bytes(raw)
                    uploaded_zips.append(p)

            if csv_files:
                for uf in csv_files:
                    if not uf.filename:
                        continue
                    raw = await uf.read()
                    if raw:
                        csv_pairs.append((uf.filename, raw))

            if not uploaded_zips and not csv_pairs:
                raise HTTPException(400, "Upload at least one ZIP or CSV file.")

            result = data_agent_run_manual(
                operator_id=operator_id.strip(),
                operator_name=operator_name.strip(),
                zip_paths=uploaded_zips if uploaded_zips else None,
                csv_pairs=csv_pairs if csv_pairs else None,
                date_range=date_range.strip(),
            )
            return JSONResponse(result)
        finally:
            shutil.rmtree(work, ignore_errors=True)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/sessions")
def api_list_sessions():
    return list_sessions()


@app.get("/api/overview")
def api_overview():
    return _build_overview()


@app.get("/api/sessions/{session_id}")
def api_get_session(session_id: str):
    try:
        return get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")


@app.post("/api/sessions/{session_id}/upload")
async def session_upload_more(
    session_id: str,
    zip_files: Optional[List[UploadFile]] = File(None),
    csv_files: Optional[List[UploadFile]] = File(None),
):
    """Add more files to an existing session."""
    try:
        get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")

    work = Path(tempfile.mkdtemp(prefix="session_upload_"))
    try:
        if zip_files:
            paths = []
            for uf in zip_files:
                if not uf.filename:
                    continue
                raw = await uf.read()
                if raw:
                    p = work / Path(uf.filename).name
                    p.write_bytes(raw)
                    paths.append(p)
            if paths:
                ingest_zip_files(session_id, paths)

        if csv_files:
            pairs = []
            for uf in csv_files:
                if not uf.filename:
                    continue
                raw = await uf.read()
                if raw:
                    pairs.append((uf.filename, raw))
            if pairs:
                ingest_csv_files(session_id, pairs)

        return JSONResponse(get_session(session_id))
    finally:
        shutil.rmtree(work, ignore_errors=True)


# ── Session-aware agent endpoints ────────────────────────────────────────────

async def _parse_operator_id(request: Request, default: str = "") -> str:
    """Extract operator_id from JSON body or form data (supports both)."""
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        try:
            body = await request.json()
            return body.get("operator_id", default)
        except Exception:
            return default
    try:
        form = await request.form()
        return form.get("operator_id", default)
    except Exception:
        return default


def _deepdive_summary_from_result(result: dict) -> str:
    sections = result.get("sections") or {}
    exec_summary = sections.get("executive_summary") or {}
    insights = exec_summary.get("insights") or []
    if insights:
        return "\n".join(f"• {i}" for i in insights)
    return "DeepDive analysis complete."


def _deepdive_report_url(run_id: str) -> str:
    path = f"/api/runs/deepdive/{run_id}/report"
    if PUBLIC_API_BASE_URL:
        return f"{PUBLIC_API_BASE_URL}{path}"
    return path


@app.post("/api/sessions/{session_id}/run/deepdive")
async def session_run_deepdive(session_id: str, request: Request):
    """Run DeepDive using data from an existing session."""
    try:
        meta = get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")

    raw_op = await _parse_operator_id(request)
    op_id = raw_op.strip() or meta["operator_id"]
    t0 = datetime.now(timezone.utc)
    try:
        data_dir = get_session_data_dir(session_id)

        # Unrecognized uploads → basic generic analysis instead of a hard error.
        if not meta.get("datasets"):
            from shared.generic_analysis import analyze_unrecognized_files

            generic = analyze_unrecognized_files(data_dir)
            if generic:
                record_agent_run(session_id, "deepdive", session_id, {
                    "status": "generic_analysis",
                    "files_analyzed": generic.get("files_analyzed", 0),
                })
                return JSONResponse({
                    "session_id": session_id,
                    **generic,
                    "message": "Uploaded file(s) are not standard delivery-platform exports — basic analysis generated.",
                })

        result = run_deepdive(operator_id=op_id, data_dir=data_dir)
        if result.get("status") != "success":
            raise HTTPException(400, result.get("message", "DeepDive failed"))

        run_id = str(uuid.uuid4())
        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()

        out_dir = DD_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        report_path = Path(result["report_html_path"])
        shutil.copy(report_path, out_dir / "report.html")

        dd_json_path = result.get("deepdive_json_path")
        if dd_json_path and Path(dd_json_path).is_file():
            shutil.copy(dd_json_path, out_dir / "deepdive.json")

        record_agent_run(session_id, "deepdive", run_id, {
            "status": "success",
            "datasets_loaded": result.get("datasets_loaded", []),
        })

        _append_index({
            "id": run_id, "agent": "deepdive", "operator": op_id,
            "status": "success",
            "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
            "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            "session_id": session_id,
        })

        report_url = _deepdive_report_url(run_id)
        return JSONResponse({
            **result,
            "run_id": run_id,
            "session_id": session_id,
            "summary": _deepdive_summary_from_result(result),
            "report_url": report_url,
            "downloads": {"HTML Report": report_url},
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/sessions/{session_id}/run/marketingreco")
async def session_run_marketingreco(session_id: str, request: Request):
    """Run MarketingReco using data from an existing session."""
    try:
        meta = get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")

    raw_op = await _parse_operator_id(request)
    op_id = raw_op.strip() or meta["operator_id"]
    t0 = datetime.now(timezone.utc)
    try:
        data_dir = get_session_data_dir(session_id)

        # Look for uploaded slot-table CSVs (downloaded from DeepDive)
        aov_csv = None
        prof_csv = None
        for p in sorted(data_dir.rglob("*.csv")):
            name_lower = p.name.lower()
            if "aov" in name_lower:
                aov_csv = str(p)
            elif "profitability" in name_lower:
                prof_csv = str(p)

        # Also look for the financial detailed CSV (builds full slot tables)
        fin_csv = None
        for p in sorted(data_dir.rglob("*FINANCIAL*DETAILED*.csv")):
            fin_csv = str(p)
            break
        if not fin_csv:
            for p in sorted(data_dir.rglob("*.csv")):
                if "financial" in p.name.lower():
                    fin_csv = str(p)
                    break

        if not aov_csv and not fin_csv:
            available = [p.name for p in data_dir.rglob("*.csv")]
            _api_log.warning("No usable CSV found. Available: %s", available)
            raise HTTPException(400, f"No slot table or financial CSV found. Upload the AOV + Profitability CSVs from DeepDive, or a FINANCIAL_DETAILED export. Available: {available}")

        result = run_marketingreco(
            op_id,
            mode="manual",
            aov_csv_path=aov_csv,
            profitability_csv_path=prof_csv,
            financial_report_path=fin_csv,
            reporting_root=str(ROOT / "agents/resgro-browser-automation"),
        )

        run_id = str(uuid.uuid4())
        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        record_agent_run(session_id, "marketingreco", run_id, {
            "status": "success",
            "campaigns": len(result.get("recommended_campaigns", [])),
        })

        _append_index({
            "id": run_id, "agent": "marketingreco", "operator": op_id,
            "status": "success",
            "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
            "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            "session_id": session_id,
        })

        return JSONResponse({**result, "run_id": run_id, "session_id": session_id})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/sessions/{session_id}/run/campaign-review")
async def session_run_campaign_review(session_id: str, request: Request):
    """Run Campaign Review using data from an existing session."""
    try:
        meta = get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")

    raw_op = await _parse_operator_id(request)
    op_id = raw_op.strip() or meta["operator_id"]
    t0 = datetime.now(timezone.utc)
    try:
        data_dir = get_session_data_dir(session_id)
        marketing_csvs = sorted(data_dir.rglob("MARKETING_*.csv"))
        data_files = [str(p) for p in marketing_csvs] if marketing_csvs else None

        result = campaign_review_to_json_safe(
            run_campaign_review(
                operator_id=op_id,
                mode="manual" if data_files else "auto",
                data_files=data_files,
            )
        )

        run_id = str(uuid.uuid4())
        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        record_agent_run(session_id, "campaign_review", run_id, {
            "status": "success",
            "reviews": len(result.get("campaign_reviews", [])),
        })

        _append_index({
            "id": run_id, "agent": "campaign_review", "operator": op_id,
            "status": "success",
            "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
            "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            "session_id": session_id,
        })

        return JSONResponse({**result, "run_id": run_id, "session_id": session_id})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/sessions/{session_id}/run/monthly-reporter")
async def session_run_monthly_reporter(
    session_id: str,
    pre_range: str = Form(...),
    post_range: str = Form(...),
    operator_id: str = Form(""),
    excluded_dates: str = Form(""),
    dd_store_ids: str = Form(""),
    ue_store_ids: str = Form(""),
):
    """Run Monthly Reporter using data from an existing session."""
    try:
        meta = get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")

    op_id = operator_id.strip() or meta["operator_id"]
    t0 = datetime.now(timezone.utc)
    run_id = str(uuid.uuid4())

    try:
        data_dir = get_session_data_dir(session_id)

        inputs = ReportInputs(
            pre_range=pre_range.strip(),
            post_range=post_range.strip(),
            excluded_dates_text=excluded_dates.strip(),
            operator_name=meta.get("operator_name", ""),
            dd_store_ids_text=dd_store_ids.strip(),
            ue_store_ids_text=ue_store_ids.strip(),
        )

        bundle = generate_monthly_report_bundle(inputs, data_root=data_dir)

        out_dir = RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        full_name = bundle["filename"]
        (out_dir / full_name).write_bytes(bundle["excel_bytes"])

        date_name = bundle.get("date_export_filename")
        if bundle.get("date_export_bytes") and date_name:
            (out_dir / date_name).write_bytes(bundle["date_export_bytes"])

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        record_agent_run(session_id, "monthly_reporter", run_id, {
            "status": "success",
            "summary": bundle.get("summary_text", ""),
        })

        _append_index({
            "id": run_id, "agent": "monthly_reporter", "operator": op_id,
            "status": "success",
            "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
            "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            "session_id": session_id,
        })

        return JSONResponse({
            "run_id": run_id,
            "session_id": session_id,
            "summary_text": bundle.get("summary_text"),
            "downloads": {
                "full": f"/api/runs/{run_id}/download/full",
                "date": f"/api/runs/{run_id}/download/date" if bundle.get("date_export_bytes") else None,
            },
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Campaign Setup endpoint ───────────────────────────────────────────────────

CS_RUNS_BASE = ROOT / "data" / "runs" / "campaign_setup"
CS_RUNS_BASE.mkdir(parents=True, exist_ok=True)


@app.post("/api/sessions/{session_id}/run/campaign-setup")
async def session_run_campaign_setup(
    session_id: str,
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    """Run Campaign Setup — create promo offers + sponsored listing campaigns via browser automation."""
    try:
        meta = get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")

    if not doordash_email.strip() or not doordash_password:
        raise HTTPException(400, "DoorDash email and password are required for campaign setup.")

    operator_id = meta["operator_id"]
    t0 = datetime.now(timezone.utc)
    run_id = str(uuid.uuid4())

    try:
        from agents.boss_agent.agent import _run_campaign_setup

        import asyncio
        result = await asyncio.to_thread(
            _run_campaign_setup,
            session_id,
            operator_id,
            doordash_email.strip(),
            doordash_password,
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        record_agent_run(session_id, "campaign_setup", run_id, {
            "status": result.get("status", "unknown"),
            "offers_status": result.get("offers", {}).get("status", "skipped"),
            "ads_status": result.get("ads", {}).get("status", "skipped"),
        })

        _append_index({
            "id": run_id, "agent": "campaign_setup", "operator": operator_id,
            "status": result.get("status", "success"),
            "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
            "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            "session_id": session_id,
        })

        return JSONResponse({**result, "run_id": run_id, "session_id": session_id})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Boss Agent endpoint ──────────────────────────────────────────────────────

@app.post("/api/run/boss")
async def run_boss_full(
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
    date_range: str = Form(""),
):
    """Run Boss Agent full pipeline: Data Agent → DeepDive → Marketing Reco → Campaign Setup."""
    if not doordash_email or not doordash_password:
        raise HTTPException(400, "DoorDash email and password are required")
    try:
        result = boss_agent_run(
            doordash_email=doordash_email.strip(),
            doordash_password=doordash_password,
            date_range=date_range.strip(),
        )
        return JSONResponse(result)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/sessions/{session_id}/run/boss")
async def session_run_boss(
    session_id: str,
    steps: str = Form(""),
    skip_steps: str = Form(""),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
    pre_range: str = Form(""),
    post_range: str = Form(""),
):
    """Run the Boss Agent against an existing session."""
    try:
        get_session(session_id)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")

    try:
        step_list = [s.strip() for s in steps.split(",") if s.strip()] or None
        skip_list = [s.strip() for s in skip_steps.split(",") if s.strip()] or None

        result = boss_agent_run(
            session_id=session_id,
            steps=step_list,
            skip_steps=skip_list,
            doordash_email=doordash_email.strip(),
            doordash_password=doordash_password,
            pre_range=pre_range.strip(),
            post_range=post_range.strip(),
        )
        return JSONResponse(result)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/runs")
def list_runs() -> list[dict]:
    return _read_all_runs()


@app.get("/api/runs/{run_id}")
def get_run(run_id: str) -> dict:
    # Try monthly_reporter first
    meta_path = RUNS_BASE / run_id / "meta.json"
    if not meta_path.is_file():
        # Try deepdive
        meta_path = DD_RUNS_BASE / run_id / "meta.json"
        
    if not meta_path.is_file():
        raise HTTPException(404, "Run not found")
    return json.loads(meta_path.read_text(encoding="utf-8"))


@app.get("/api/runs/{run_id}/preview")
def get_preview(run_id: str) -> dict:
    p = RUNS_BASE / run_id / "preview.json"
    if not p.is_file():
        raise HTTPException(404, "Preview not found")
    return json.loads(p.read_text(encoding="utf-8"))


@app.get("/api/runs/{run_id}/download/full")
def download_full(run_id: str):
    folder = RUNS_BASE / run_id
    meta_path = folder / "meta.json"
    if not meta_path.is_file():
        raise HTTPException(404, "Run not found")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    fn = meta.get("full_report_filename") or "report.xlsx"
    path = folder / fn
    if not path.is_file():
        raise HTTPException(404, "File missing")
    return FileResponse(path, filename=fn, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/runs/{run_id}/download/date")
def download_date(run_id: str):
    folder = RUNS_BASE / run_id
    meta_path = folder / "meta.json"
    if not meta_path.is_file():
        raise HTTPException(404, "Run not found")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    fn = meta.get("date_export_filename")
    if not fn:
        raise HTTPException(404, "Date export not available for this run")
    path = folder / fn
    if not path.is_file():
        raise HTTPException(404, "File missing")
    return FileResponse(path, filename=fn, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/runs/deepdive")
async def post_deepdive(
    operator_id: str = Form(...),
    zip_files: Optional[List[UploadFile]] = File(
        None,
        description="Required: one or more DoorDash export zip files for DeepDive.",
    ),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"dd_{run_id[:8]}_"))
    try:
        uploaded_names: list[str] = []
        if not zip_files:
            raise HTTPException(400, "Upload at least one DeepDive zip file.")

        for uploaded in zip_files:
            filename = (uploaded.filename or "").strip()
            if not filename:
                continue
            if not filename.lower().endswith(".zip"):
                raise HTTPException(400, f"Invalid file '{filename}'. Only .zip files are allowed.")
            raw = await uploaded.read()
            if not raw:
                continue
            safe_name = Path(filename).name
            (work / safe_name).write_bytes(raw)
            uploaded_names.append(safe_name)

        if not uploaded_names:
            raise HTTPException(400, "Upload at least one non-empty DeepDive zip file.")

        res = run_deepdive(operator_id=operator_id, data_dir=work)
        if res.get("status") != "success":
            raise HTTPException(400, res.get("message", "DeepDive failed"))

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()

        report_path = Path(res["report_html_path"])

        # Store in DD_RUNS_BASE / run_id
        out_dir = DD_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)

        shutil.copy(report_path, out_dir / "report.html")

        legacy_path_str = res.get("deepdive_json_path") or ""
        legacy_path = Path(legacy_path_str) if legacy_path_str else None
        deepdive_report: dict | None = None
        if legacy_path and legacy_path.is_file():
            deepdive_report = json.loads(legacy_path.read_text(encoding="utf-8"))
            shutil.copy(legacy_path, out_dir / "deepdive.json")

        meta = {
            "run_id": run_id,
            "agent": "deepdive",
            "operator_id": operator_id,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "uploaded_files": uploaded_names,
            "datasets_loaded": res.get("datasets_loaded", []),
            "upload_audit": res.get("upload_audit", {}),
            "metric_hierarchy": (res.get("sections") or {}).get("metric_hierarchy") or {},
            "deepdive_json_path": res.get("deepdive_json_path"),
            "report_url": (
                f"{PUBLIC_API_BASE_URL}/api/runs/deepdive/{run_id}/report"
                if PUBLIC_API_BASE_URL
                else f"/api/runs/deepdive/{run_id}/report"
            ),
        }
        if deepdive_report is not None:
            meta["deepdive_report"] = deepdive_report
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

        _append_index(
            {
                "id": run_id,
                "agent": "deepdive",
                "operator": operator_id,
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )

        return JSONResponse(meta)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.get("/api/runs/deepdive/{run_id}/report")
def get_deepdive_report(run_id: str):
    path = DD_RUNS_BASE / run_id / "report.html"
    if not path.is_file():
        raise HTTPException(404, "Report not found")
    return FileResponse(path)


@app.post("/api/runs/marketingreco")
async def post_marketingreco(
    operator_id: str = Form(...),
    mode: str = Form("manual"),
    financial_file: Optional[UploadFile] = File(None),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"mrk_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        kwargs: dict = {}

        if mode_norm == "manual":
            if not financial_file or not financial_file.filename:
                raise HTTPException(400, "Manual mode requires FINANCIAL_DETAILED file (.zip or .csv).")
            if not (
                financial_file.filename.lower().endswith(".zip")
                or financial_file.filename.lower().endswith(".csv")
            ):
                raise HTTPException(400, "financial_file must be .zip or .csv")
            raw = await financial_file.read()
            if not raw:
                raise HTTPException(400, "financial_file is empty.")
            in_path = work / Path(financial_file.filename).name
            in_path.write_bytes(raw)
            kwargs["financial_report_path"] = str(in_path)
            kwargs["reporting_root"] = str(ROOT / "agents/resgro-browser-automation")
        elif mode_norm == "auto":
            if not doordash_email.strip() or not doordash_password:
                raise HTTPException(400, "Auto mode requires doordash_email and doordash_password.")
            kwargs["doordash_email"] = doordash_email.strip()
            kwargs["doordash_password"] = doordash_password
            kwargs["reporting_root"] = str(ROOT / "agents/resgro-browser-automation")
        else:
            raise HTTPException(400, "mode must be 'manual' or 'auto'")

        result = run_marketingreco(operator_id=operator_id.strip(), mode=mode_norm, **kwargs)

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        out_dir = MRK_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "run_id": run_id,
            "agent": "marketingreco",
            "operator_id": operator_id.strip(),
            "mode": mode_norm,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "recommended_campaigns": len(result.get("recommended_campaigns") or []),
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
        (out_dir / "result.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
        campaigns_xlsx = out_dir / "marketingreco_campaigns.xlsx"
        _write_marketingreco_campaigns_excel(campaigns_xlsx, result)
        _append_index(
            {
                "id": run_id,
                "agent": "marketingreco",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        ads_plan_payload = result.get("ads_plan") or {}
        response = {
            **result,
            "run_id": run_id,
            "ads_upload_rows": resgro_ads_upload_rows(ads_plan_payload),
            "downloads": {
                "campaigns_excel": f"/api/runs/marketingreco/{run_id}/download/campaigns",
            },
        }
        return JSONResponse(response)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


class _MarketingRecoFromGcsRequest(BaseModel):
    operator_id: str
    mode: str = "manual"
    financial_object: _GcsObjectRef
    doordash_email: str = ""
    doordash_password: str = ""


@app.post("/api/runs/marketingreco/from-gcs")
def post_marketingreco_from_gcs(req: _MarketingRecoFromGcsRequest):
    if not _gcs.uploads_enabled():
        raise HTTPException(503, "GCS uploads are not configured on this server.")

    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"mrk_gcs_{run_id[:8]}_"))
    gcs_paths: list[str] = []

    try:
        mode_norm = req.mode.strip().lower()
        if mode_norm != "manual":
            raise HTTPException(400, "from-gcs only supports manual mode with financial_object.")

        in_path = work / Path(req.financial_object.filename).name
        _gcs.download_object_to_path(req.financial_object.object_path, in_path)
        gcs_paths.append(req.financial_object.object_path)
        if not in_path.stat().st_size:
            raise HTTPException(400, "financial file is empty.")

        result = run_marketingreco(
            operator_id=req.operator_id.strip(),
            mode="manual",
            financial_report_path=str(in_path),
            reporting_root=str(ROOT / "agents/resgro-browser-automation"),
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        out_dir = MRK_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "run_id": run_id,
            "agent": "marketingreco",
            "operator_id": req.operator_id.strip(),
            "mode": mode_norm,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "recommended_campaigns": len(result.get("recommended_campaigns") or []),
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
        (out_dir / "result.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
        campaigns_xlsx = out_dir / "marketingreco_campaigns.xlsx"
        _write_marketingreco_campaigns_excel(campaigns_xlsx, result)
        _append_index(
            {
                "id": run_id,
                "agent": "marketingreco",
                "operator": req.operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        ads_plan_payload = result.get("ads_plan") or {}
        return JSONResponse(
            {
                **result,
                "run_id": run_id,
                "ads_upload_rows": resgro_ads_upload_rows(ads_plan_payload),
                "downloads": {
                    "campaigns_excel": f"/api/runs/marketingreco/{run_id}/download/campaigns",
                },
            }
        )
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(500, str(exc)) from exc
    finally:
        shutil.rmtree(work, ignore_errors=True)
        _gcs.delete_objects(gcs_paths)


class _DeepDiveFromGcsRequest(BaseModel):
    operator_id: str
    objects: list[_GcsObjectRef]


@app.post("/api/runs/deepdive/from-gcs")
def post_deepdive_from_gcs(req: _DeepDiveFromGcsRequest):
    if not req.objects:
        raise HTTPException(400, "Provide at least one uploaded zip object.")
    if not _gcs.uploads_enabled():
        raise HTTPException(503, "GCS uploads are not configured on this server.")

    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"dd_gcs_{run_id[:8]}_"))
    gcs_paths: list[str] = []

    try:
        uploaded_zips, _csv_pairs, gcs_paths = _gcs.materialize_objects(
            [o.model_dump() for o in req.objects],
            work,
        )
        if not uploaded_zips:
            raise HTTPException(400, "DeepDive from-gcs requires at least one .zip file.")

        res = run_deepdive(operator_id=req.operator_id.strip(), data_dir=work)
        if res.get("status") != "success":
            raise HTTPException(400, res.get("message", "DeepDive failed"))

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        report_path = Path(res["report_html_path"])
        out_dir = DD_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy(report_path, out_dir / "report.html")

        legacy_path_str = res.get("deepdive_json_path") or ""
        legacy_path = Path(legacy_path_str) if legacy_path_str else None
        deepdive_report: dict | None = None
        if legacy_path and legacy_path.is_file():
            deepdive_report = json.loads(legacy_path.read_text(encoding="utf-8"))
            shutil.copy(legacy_path, out_dir / "deepdive.json")

        uploaded_names = [p.name for p in uploaded_zips]
        meta = {
            "run_id": run_id,
            "agent": "deepdive",
            "operator_id": req.operator_id.strip(),
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "uploaded_files": uploaded_names,
            "datasets_loaded": res.get("datasets_loaded", []),
            "upload_audit": res.get("upload_audit", {}),
            "metric_hierarchy": (res.get("sections") or {}).get("metric_hierarchy") or {},
            "deepdive_json_path": res.get("deepdive_json_path"),
            "report_url": _deepdive_report_url(run_id),
        }
        if deepdive_report is not None:
            meta["deepdive_report"] = deepdive_report
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

        _append_index(
            {
                "id": run_id,
                "agent": "deepdive",
                "operator": req.operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        return JSONResponse(meta)
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(500, str(exc)) from exc
    finally:
        shutil.rmtree(work, ignore_errors=True)
        _gcs.delete_objects(gcs_paths)


class _CampaignReviewFromGcsRequest(BaseModel):
    operator_id: str
    mode: str = "manual"
    data_dir: str = ""
    objects: list[_GcsObjectRef]


@app.post("/api/runs/campaign-review/from-gcs")
def post_campaign_review_from_gcs(req: _CampaignReviewFromGcsRequest):
    if not req.objects:
        raise HTTPException(400, "Provide at least one uploaded marketing file.")
    if not _gcs.uploads_enabled():
        raise HTTPException(503, "GCS uploads are not configured on this server.")

    mode_norm = req.mode.strip().lower()
    if mode_norm != "manual":
        raise HTTPException(400, "from-gcs only supports manual mode with marketing files.")

    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"cr_gcs_{run_id[:8]}_"))
    gcs_paths: list[str] = []

    try:
        data_files: list[str] = []
        for obj in req.objects:
            local = work / Path(obj.filename).name
            _gcs.download_object_to_path(obj.object_path, local)
            gcs_paths.append(obj.object_path)
            if local.stat().st_size:
                data_files.append(str(local))
        if not data_files:
            raise HTTPException(400, "No non-empty marketing files in GCS upload.")

        result = campaign_review_to_json_safe(
            run_campaign_review(
                operator_id=req.operator_id.strip(),
                mode="manual",
                data_dir=(req.data_dir.strip() or None),
                data_files=data_files,
            )
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        out_dir = CR_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "run_id": run_id,
            "agent": "campaign_review",
            "operator_id": req.operator_id.strip(),
            "mode": mode_norm,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "campaign_reviews": len(result.get("campaign_reviews") or []),
            "datasets_loaded": (result.get("summary_metrics") or {}).get("datasets_loaded", []),
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
        (out_dir / "result.json").write_text(
            json.dumps(result, indent=2, allow_nan=False),
            encoding="utf-8",
        )
        _append_index(
            {
                "id": run_id,
                "agent": "campaign_review",
                "operator": req.operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        return JSONResponse({**result, "run_id": run_id})
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(500, str(exc)) from exc
    finally:
        shutil.rmtree(work, ignore_errors=True)
        _gcs.delete_objects(gcs_paths)


@app.post("/api/runs/offers")
async def post_offers(
    operator_id: str = Form(...),
    mode: str = Form("manual"),
    campaign_mappings_file: Optional[UploadFile] = File(None),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"offers_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        reporting_root = ROOT / "agents/resgro-browser-automation"
        env = os.environ.copy()
        if not doordash_email.strip() or not doordash_password:
            raise HTTPException(400, "Offers mode requires doordash_email and doordash_password.")
        env["DOORDASH_EMAIL"] = doordash_email.strip()
        env["DOORDASH_PASSWORD"] = doordash_password
        if mode_norm not in ("manual", "auto", "full"):
            raise HTTPException(400, "mode must be 'manual', 'auto', or 'full'")

        # Product behavior: Offers mode always runs the complete Reporting app pipeline
        # (download + analysis + campaign execution) with credentials provided via UI.
        _ = campaign_mappings_file
        subprocess.run(
            [sys.executable, "main.py"],
            cwd=str(reporting_root),
            env=env,
            check=True,
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "offers",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        return JSONResponse({"status": "success", "mode": "full", "operator_id": operator_id.strip()})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.post("/api/runs/ads")
async def post_ads(
    operator_id: str = Form(...),
    mode: str = Form("manual"),
    ads_sheet_file: Optional[UploadFile] = File(None),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    """
    Sponsored listing automation.

    Manual: CSV or any Excel file; for Excel, sheet "Ads" is read as input rows.
    Expected Ads columns: Merchant store ID (or Store ID) | Slots | Bid strategy | Budget | Campaign name.
    Auto: login → download financial + marketing reports → analysis + combined workbook (campaign
    recommendations) → build Ads upload rows from FINANCIAL_DETAILED → sponsored listing automation
    (same browser flow as Manual).
    """
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"ads_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        reporting_root = ROOT / "agents/resgro-browser-automation"
        if mode_norm not in ("manual", "auto"):
            raise HTTPException(400, "mode must be 'manual' or 'auto'")
        if not doordash_email.strip() or not doordash_password:
            raise HTTPException(400, "DoorDash email and password are required (browser login).")

        env = os.environ.copy()
        env["DOORDASH_EMAIL"] = doordash_email.strip()
        env["DOORDASH_PASSWORD"] = doordash_password

        rows_file: str | None = None

        if mode_norm == "manual":
            if not ads_sheet_file or not ads_sheet_file.filename:
                raise HTTPException(400, "Manual mode requires an ads sheet (.csv or Excel).")
            fn = ads_sheet_file.filename.lower()
            if not (
                fn.endswith(".csv")
                or fn.endswith(".xlsx")
                or fn.endswith(".xls")
                or fn.endswith(".xlsm")
                or fn.endswith(".xltx")
                or fn.endswith(".xltm")
            ):
                raise HTTPException(400, "ads_sheet_file must be .csv or an Excel file")

            raw = await ads_sheet_file.read()
            if not raw:
                raise HTTPException(400, "ads_sheet_file is empty.")

            sheet_path = work / Path(ads_sheet_file.filename).name
            sheet_path.write_bytes(raw)
            rows_path = _prepare_ads_rows_file(sheet_path, work)
            rows_file = rows_path.name

            env["ADS_DOWNLOAD_DIR"] = str(work)
            env["ADS_SHEET_PATH"] = str(rows_path)

            script = """
import asyncio
import os
from pathlib import Path
from agents.doordash_agent import run_ads_campaigns_from_sheet

async def _main():
    await run_ads_campaigns_from_sheet(
        download_dir=Path(os.environ["ADS_DOWNLOAD_DIR"]),
        email=os.environ["DOORDASH_EMAIL"],
        password=os.environ["DOORDASH_PASSWORD"],
        sheet_path=Path(os.environ["ADS_SHEET_PATH"]),
    )

asyncio.run(_main())
"""
            subprocess.run(
                [sys.executable, "-c", script],
                cwd=str(reporting_root),
                env=env,
                check=True,
            )
        elif mode_norm == "auto":
            env["RESGRO_AI_ROOT"] = str(ROOT)
            ads_auto_script = """
import asyncio
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

from agents.doordash_agent import run_ads_campaigns_from_sheet, run_reports_only
from agents.marketing_agent import run as marketing_run
from agents.analysis_agent import run as analysis_run
from agents.combined_report_agent import run as combined_run, append_campaign_mappings_to_workbook
from agents.campaign_params import get_campaign_mappings_for_combined
import pandas as pd


def _dates():
    today = datetime.now().date()
    first_this_month = today.replace(day=1)
    last_prev_month = first_this_month - timedelta(days=1)
    y, m = first_this_month.year, first_this_month.month - 3
    if m <= 0:
        m += 12
        y -= 1
    start = datetime(y, m, 1).date()
    return start.strftime("%m/%d/%Y"), last_prev_month.strftime("%m/%d/%Y")


def _run_dir(email: str) -> Path:
    safe = (email or "run").strip()
    for c in ("@", ".", " ", "/", chr(92)):
        safe = safe.replace(c, "_")
    safe = safe[:50] if len(safe) > 50 else safe
    return Path("downloads") / f"{safe}-{datetime.now().strftime('%Y%m%d_%H%M%S')}"


async def _main():
    email = os.environ["DOORDASH_EMAIL"]
    password = os.environ["DOORDASH_PASSWORD"]
    resgro_root = Path(os.environ["RESGRO_AI_ROOT"])
    sys.path.insert(0, str(resgro_root))
    from agents.marketingreco.ads_planner import build_ads_plan
    from agents.marketingreco.resgro_ads_excel import resgro_ads_upload_rows

    start_date, end_date = _dates()
    run_dir = _run_dir(email)
    run_dir.mkdir(parents=True, exist_ok=True)

    marketing_path, financial_path = await run_reports_only(
        download_dir=run_dir,
        email=email,
        password=password,
        start_date=start_date,
        end_date=end_date,
    )
    if not financial_path:
        raise SystemExit(
            "Ads auto: financial report was not downloaded. Check Browser Use / portal access."
        )

    marketing_sheets = (
        marketing_run(
            Path(marketing_path),
            output_dir=run_dir,
            post_start_date=start_date,
            post_end_date=end_date,
            write_file=False,
        )
        if marketing_path
        else None
    )
    financial_sheets = analysis_run(
        Path(financial_path),
        output_dir=run_dir,
        report_start_date=start_date,
        report_end_date=end_date,
        write_file=False,
    )

    combined = combined_run(
        financial_sheets=financial_sheets,
        marketing_sheets=marketing_sheets,
        output_dir=run_dir,
    )
    if combined:
        slots_csv = Path("slots.csv")
        mappings = get_campaign_mappings_for_combined(Path(combined), slots_csv)
        if mappings:
            append_campaign_mappings_to_workbook(Path(combined), mappings)

    fc = run_dir / "financial_detailed_report.csv"
    if not fc.is_file():
        for p in sorted(run_dir.glob("*FINANCIAL*.csv")):
            fc = p
            break
    if not fc.is_file():
        raise SystemExit(
            "Ads auto: no FINANCIAL_DETAILED CSV after analysis; cannot build ads recommendations."
        )

    ads_plan = build_ads_plan(str(fc))
    upload = resgro_ads_upload_rows(ads_plan)
    if not upload:
        raise SystemExit(
            "Ads auto: no sponsored-listing rows (no slots with Ad placement Yes). "
            "Try Manual with an Ads sheet or check financial data coverage."
        )

    ads_csv = run_dir / "ads_auto_upload.csv"
    pd.DataFrame(upload).to_csv(ads_csv, index=False)

    await run_ads_campaigns_from_sheet(
        download_dir=run_dir,
        email=email,
        password=password,
        sheet_path=ads_csv,
    )


asyncio.run(_main())
"""
            subprocess.run(
                [sys.executable, "-c", ads_auto_script],
                cwd=str(reporting_root),
                env=env,
                check=True,
            )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "ads",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        body: dict = {
            "status": "success",
            "run_id": run_id,
            "mode": mode_norm,
            "operator_id": operator_id.strip(),
        }
        if rows_file:
            body["rows_file"] = rows_file
        return JSONResponse(body)
    except HTTPException:
        raise
    except subprocess.CalledProcessError as e:
        raise HTTPException(500, f"Ads browser run failed (exit {e.returncode}). Check API logs / Slack.")
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.get("/api/runs/marketingreco/{run_id}/download/campaigns")
def download_marketingreco_campaigns(run_id: str):
    path = MRK_RUNS_BASE / run_id / "marketingreco_campaigns.xlsx"
    if not path.is_file():
        raise HTTPException(404, "Campaign table not found")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/runs/campaign-review")
async def post_campaign_review(
    operator_id: str = Form(...),
    mode: str = Form("auto"),
    marketing_files: Optional[List[UploadFile]] = File(None),
    data_dir: str = Form(""),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"cr_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        if mode_norm not in ("auto", "manual"):
            raise HTTPException(400, "mode must be 'auto' or 'manual'")

        data_files: list[str] = []
        if mode_norm == "manual":
            if not marketing_files:
                raise HTTPException(
                    400,
                    "Manual mode requires marketing_files (MARKETING_PROMOTION* / MARKETING_SPONSORED_LISTING* csv/zip).",
                )
            for uf in marketing_files:
                if not uf.filename:
                    continue
                raw = await uf.read()
                if not raw:
                    continue
                p = work / Path(uf.filename).name
                p.write_bytes(raw)
                data_files.append(str(p))
            if not data_files:
                raise HTTPException(400, "No non-empty files uploaded for manual campaign review.")

        result = campaign_review_to_json_safe(
            run_campaign_review(
                operator_id=operator_id.strip(),
                mode=mode_norm,  # type: ignore[arg-type]
                data_dir=(data_dir.strip() or None),
                data_files=data_files if mode_norm == "manual" else None,
            )
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        out_dir = CR_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "run_id": run_id,
            "agent": "campaign_review",
            "operator_id": operator_id.strip(),
            "mode": mode_norm,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "campaign_reviews": len(result.get("campaign_reviews") or []),
            "datasets_loaded": (result.get("summary_metrics") or {}).get("datasets_loaded", []),
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
        (out_dir / "result.json").write_text(
            json.dumps(result, indent=2, allow_nan=False),
            encoding="utf-8",
        )
        _append_index(
            {
                "id": run_id,
                "agent": "campaign_review",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        return JSONResponse({**result, "run_id": run_id})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.post("/api/runs/monthly-reporter")
async def post_monthly_reporter(
    pre_range: str = Form(..., description="MM/DD/YYYY-MM/DD/YYYY"),
    post_range: str = Form(...),
    operator_id: str = Form(""),
    operator_name: str = Form(""),
    excluded_dates: str = Form(""),
    dd_store_ids: str = Form(""),
    ue_store_ids: str = Form(""),
    dd_file: Optional[UploadFile] = File(
        None,
        description="Optional DoorDash financial CSV (saved as dd-data.csv when provided).",
    ),
    ue_file: Optional[UploadFile] = File(
        None,
        description="Optional UberEats financial CSV (saved as ue-data.csv when provided).",
    ),
    marketing_files: Optional[List[UploadFile]] = File(
        None,
        description="Optional: multiple MARKETING_*.csv (Streamlit file_upload_screen behavior)",
    ),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"mr_{run_id[:8]}_"))

    try:
        # Streamlit parity: financial + marketing files are optional; only Pre/Post dates are required.
        if dd_file and dd_file.filename:
            raw = await dd_file.read()
            if raw:
                (work / "dd-data.csv").write_bytes(raw)
        if ue_file and ue_file.filename:
            raw = await ue_file.read()
            if raw:
                (work / "ue-data.csv").write_bytes(raw)

        # Marketing CSVs — same layout as Streamlit `file_upload_screen` (marketing_data/marketing_*).
        mkt_pairs: list[tuple[str, bytes]] = []
        if marketing_files:
            for uf in marketing_files:
                if uf.filename:
                    raw = await uf.read()
                    if raw:
                        mkt_pairs.append((uf.filename, raw))
        if mkt_pairs:
            write_marketing_csvs_to_work_dir(work, mkt_pairs)

        inputs = ReportInputs(
            pre_range=pre_range.strip(),
            post_range=post_range.strip(),
            excluded_dates_text=excluded_dates.strip(),
            operator_name=operator_name.strip(),
            dd_store_ids_text=dd_store_ids.strip(),
            ue_store_ids_text=ue_store_ids.strip(),
        )

        bundle = generate_monthly_report_bundle(inputs, data_root=work)

        out_dir = RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)

        full_name = bundle["filename"]
        (out_dir / full_name).write_bytes(bundle["excel_bytes"])

        date_name = bundle.get("date_export_filename")
        if bundle.get("date_export_bytes") and date_name:
            (out_dir / date_name).write_bytes(bundle["date_export_bytes"])  # type: ignore[index]

        preview = {"tables": bundle.get("tables") or {}, "summary_text": bundle.get("summary_text")}
        (out_dir / "preview.json").write_text(json.dumps(preview, default=str), encoding="utf-8")

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()

        meta = {
            "run_id": run_id,
            "agent": "monthly_reporter",
            "operator_id": operator_id.strip() or "—",
            "operator_name": operator_name.strip(),
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "summary_text": bundle.get("summary_text"),
            "full_report_filename": full_name,
            "date_export_filename": date_name if bundle.get("date_export_bytes") else None,
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

        _append_index(
            {
                "id": run_id,
                "agent": "monthly_reporter",
                "operator": operator_id.strip() or operator_name.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )

        return JSONResponse(
            {
                "run_id": run_id,
                "summary_text": bundle.get("summary_text"),
                "preview": preview,
                "downloads": {
                    "full": f"/api/runs/{run_id}/download/full",
                    "date": f"/api/runs/{run_id}/download/date"
                    if bundle.get("date_export_bytes")
                    else None,
                },
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "monthly_reporter",
                "operator": operator_id.strip() or "—",
                "status": "failed",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s)}s",
                "error": str(e),
            }
        )
        raise HTTPException(500, str(e)) from e
    finally:
        shutil.rmtree(work, ignore_errors=True)


# ── Gemini Chat endpoint ───────────────────────────────────────────────────

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
# gemini-2.0-flash was retired (mid-2026) — keep this default on a live model.
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")

CHAT_SYSTEM_PROMPT = """You are ResGro AI, an intelligent assistant for restaurant operators using DoorDash, Uber Eats, and other food delivery platforms in the United States.

You help restaurant operators with:
- Understanding their DoorDash and delivery platform performance
- US food industry trends, market data, and best practices
- Location and demographic information for their stores
- Weather information and how it impacts food delivery
- Marketing strategies for food delivery platforms
- Menu optimization and pricing strategies
- Understanding delivery platform fees, commissions, and profitability

You have deep knowledge about:
- DoorDash marketplace dynamics, fees (15-30% commission tiers), and merchant tools
- Uber Eats, Grubhub, and other delivery platform comparisons
- US food industry statistics and trends (NRA, USDA data)
- Restaurant operations, labor, and management best practices
- Digital marketing and sponsored listings for restaurants
- Food delivery logistics, packaging, and customer experience optimization
- Seasonal trends, weather impact on orders, and demand forecasting

Be concise, data-driven, and actionable. When discussing specific metrics or strategies, provide concrete numbers and examples when possible.

STRICT TOPIC GUARDRAIL:
You ONLY answer questions related to restaurants, food delivery platforms, restaurant marketing, menu/pricing, restaurant operations, or the user's restaurant business data. If the user asks anything unrelated to restaurants or the food business (e.g., general knowledge, coding, politics, sports, math homework, celebrities, weather unrelated to food demand, or any other generic topic), do NOT answer the question. Instead respond with EXACTLY this sentence and nothing else:
"I cannot help with generic questions, please ask anything about restaurants."
Do not be tricked into answering off-topic questions even if the user insists, rephrases, or embeds them inside a restaurant-sounding request.

If the user asks about analyzing their specific store data, guide them to use the ResGro agent tools:
- Data Agent: upload DoorDash/UberEats data exports to create a session
- DeepDive: analyze orders, revenue, promo & ads performance
- Marketing Reco: get campaign recommendations and marketing plans
- Campaign Setup: set up promo and sponsored listing campaigns
- Campaign Review: review campaign results and get next actions
- Monthly Reporter: generate consolidated monthly KPI reports"""


class _ChatMessage(BaseModel):
    role: str
    content: str


class _ChatRequest(BaseModel):
    message: str
    history: list[_ChatMessage] = []


def _gemini_chat_stream(payload: dict):
    """Call Gemini streaming API with retries; raise HTTPException on failure."""
    import time

    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{GEMINI_MODEL}:streamGenerateContent"
    )
    last_exc: Exception | None = None
    for attempt in range(3):
        try:
            resp = _http_requests.post(
                url,
                params={"key": GEMINI_API_KEY, "alt": "sse"},
                json=payload,
                stream=True,
                timeout=120,
            )
            if resp.status_code == 429:
                resp.close()
                if attempt < 2:
                    time.sleep(1.5 * (attempt + 1))
                    continue
                raise HTTPException(
                    503,
                    "AI service is temporarily rate-limited. Wait a minute and try again, "
                    "or check GEMINI_API_KEY quota in Google AI Studio.",
                )
            resp.raise_for_status()
            return resp
        except HTTPException:
            raise
        except _http_requests.RequestException as exc:
            last_exc = exc
            if attempt < 2:
                time.sleep(1.0)
                continue
            break

    detail = "AI service unavailable."
    if last_exc is not None and last_exc.response is not None:
        status = last_exc.response.status_code
        if status == 429:
            raise HTTPException(
                503,
                "AI service is temporarily rate-limited. Wait a minute and try again, "
                "or check GEMINI_API_KEY quota in Google AI Studio.",
            ) from last_exc
        if status in (401, 403):
            raise HTTPException(
                503,
                "AI service authentication failed. Check GEMINI_API_KEY on Cloud Run.",
            ) from last_exc
    raise HTTPException(503, detail) from last_exc


@app.post("/api/chat")
def chat_gemini(req: _ChatRequest):
    if not GEMINI_API_KEY:
        raise HTTPException(500, "GEMINI_API_KEY not configured. Set the GEMINI_API_KEY environment variable.")

    contents = []
    for m in req.history:
        contents.append({"role": m.role, "parts": [{"text": m.content}]})
    contents.append({"role": "user", "parts": [{"text": req.message}]})

    payload = {
        "contents": contents,
        "systemInstruction": {"parts": [{"text": CHAT_SYSTEM_PROMPT}]},
        "generationConfig": {"temperature": 0.7, "maxOutputTokens": 4096},
    }

    resp = _gemini_chat_stream(payload)

    def _stream():
        try:
            for line in resp.iter_lines(decode_unicode=True):
                if not line or not line.startswith("data: "):
                    continue
                raw = line[6:].strip()
                if not raw:
                    continue
                try:
                    data = json.loads(raw)
                    parts = data.get("candidates", [{}])[0].get("content", {}).get("parts", [])
                    for part in parts:
                        text = part.get("text", "")
                        if text:
                            yield f"data: {json.dumps({'text': text})}\n\n"
                except (json.JSONDecodeError, KeyError, IndexError):
                    continue
            yield "data: [DONE]\n\n"
        finally:
            resp.close()

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )
